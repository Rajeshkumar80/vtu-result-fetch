# === VTU SCRAPER WITH SUBJECT EXTRACTION & URL ARG SUPPORT (FIXED) ===
# This file is a patched version of your uploaded bulk_fetcher_6.py.
# Fixes: robust alert handling, safe refresh, stable CAPTCHA retry loop.

import os
import re
import sys
import time
import traceback
import cv2
import numpy as np
import tensorflow as tf
import pandas as pd
from bs4 import BeautifulSoup
from tensorflow.keras import backend as K
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoAlertPresentException, UnexpectedAlertPresentException, WebDriverException
)

from excel_header import HEADER_ROWS, DATA_HEADER_ROW, apply_header

# -----------------------------------------
# Config
# -----------------------------------------
VTU_RESULTS_URL = "https://results.vtu.ac.in/JJEcbcs25/index.php"

# Accept URL from app.py if provided
if len(sys.argv) > 1:
    VTU_RESULTS_URL = sys.argv[1]

# Unique run identifier (passed by app.py) — appears in logs with [SUBJECTS]:
RUN_ID = sys.argv[2] if len(sys.argv) > 2 else time.strftime("%Y%m%d-%H%M%S")

MODEL_FILE = "vtu_captcha_predictor.h5"
INPUT_CSV = "students.csv"
RAW_DATA = "raw_results.csv"
RAW_SUMMARY = "raw_summary.csv"
OUTPUT_EXCEL = "vtu_results.xlsx"

IMG_WIDTH = 160
IMG_HEIGHT = 75
CHARACTERS = "0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
num_to_char = {i: c for i, c in enumerate(CHARACTERS)}

MAX_ATTEMPTS = 5


# -----------------------------------------
# Helpers
# -----------------------------------------
def preprocess(path):
    img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
    if img is None:
        raise ValueError(f"Image not found or unreadable: {path}")
    img = cv2.GaussianBlur(img, (5, 5), 0)
    _, img = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    img = cv2.resize(img, (IMG_WIDTH, IMG_HEIGHT))
    img = img / 255.0
    return img.reshape(1, IMG_HEIGHT, IMG_WIDTH, 1)


def decode(pred):
    inp = np.ones(pred.shape[0]) * pred.shape[1]
    res = K.ctc_decode(pred, input_length=inp, greedy=True)[0][0]
    out = ""
    for x in res[0]:
        if x == -1:
            break
        out += num_to_char.get(x.numpy(), "")
    return out[:6]


def predict_captcha(model, img_batch):
    """Faster single-image inference: direct eager call instead of model.predict().

    model.predict() has extra overhead (data adaption, step loops) that is
    significant when solving one CAPTCHA at a time. Calling the model directly
    runs a single forward pass. Falls back to predict() if anything fails.
    """
    try:
        out = model(img_batch, training=False)
        if isinstance(out, (list, tuple)):
            out = out[0]
        return out.numpy() if hasattr(out, "numpy") else np.asarray(out)
    except Exception:
        return model.predict(img_batch, verbose=0)


def scrape(html):
    soup = BeautifulSoup(html, "html.parser")

    usn, name = "UNKNOWN", "UNKNOWN"

    # --- Student identity: main layout (table-condensed) ---
    try:
        t = soup.find("table", {"class": "table-condensed"}).find_all("tr")
        usn = t[0].find_all("td")[1].text.strip().replace(":", "").strip().upper()
        name = t[1].find_all("td")[1].text.strip().replace(":", "").strip()
    except Exception:
        pass

    # --- Fallback identity: div-table layout (new VTU pages) ---
    if usn == "UNKNOWN" or name == "UNKNOWN":
        for row in soup.find_all("div", {"class": "divTableRow"}):
            cells = row.find_all("div", {"class": "divTableCell"})
            if len(cells) < 2:
                continue
            label = cells[0].get_text(strip=True)
            if "University Seat Number" in label and usn == "UNKNOWN":
                usn = cells[1].get_text(strip=True).replace(":", "").strip().upper()
            elif "Student Name" in label:
                name = cells[1].get_text(strip=True).replace(":", "").strip()

    # --- Fallback: find the USN pattern anywhere in the page ---
    if usn == "UNKNOWN" or not re.match(r"\d[A-Z]{2}\d{2}[A-Z]{2}\d{3}", usn):
        m = re.search(r"\d[A-Z]{2}\d{2}[A-Z]{2}\d{3}", html)
        if m:
            usn = m.group(0)
            print("[Info] USN recovered from page text:", usn)

    sub_rows = []
    total, max_total = 0, 0

    # --- Subject table: div-table layout (6 or 7 columns, both VTU layouts).
    # Some pages carry several divTableBody blocks (identity / subjects /
    # legend), so scan every divTableRow in the document. ---
    rows = soup.find_all("div", {"class": "divTableRow"})
    for r in rows:
        c = r.find_all("div", {"class": "divTableCell"})
        if len(c) < 6:
            continue
        code = c[0].text.strip()
        # Skip header / legend / info rows: a subject code is a short
        # alphanumeric token like BCS801, BINT803B, 20MATCS11
        if "Subject Code" in code or not re.match(r"^[A-Za-z0-9]{2,8}$", code):
            continue

        tmarks = c[4].text.strip()

        sub_rows.append({
            "Subject Code": code,
            "Subject Name": c[1].text.strip(),
            "Internal Marks": c[2].text.strip(),
            "External Marks": c[3].text.strip(),
            "Total Marks": tmarks,
            "Result": c[5].text.strip(),
            "Announced / Updated on": c[6].text.strip() if len(c) > 6 else "",
        })

        try:
            total += int(tmarks)
        except:
            pass

        max_total += 100

    # --- Fallback: any HTML table whose header row mentions "Subject Code" ---
    if not sub_rows:
        for table in soup.find_all("table"):
            header_cells = table.find_all("th")
            headers = [h.get_text(" ", strip=True) for h in header_cells]
            if not any("Subject Code" in h for h in headers):
                continue
            col = {h: i for i, h in enumerate(headers)}
            for tr in table.find_all("tr"):
                cells = tr.find_all("td")
                if len(cells) < 7 or "Subject Code" in cells[0].get_text(strip=True):
                    continue
                cell = lambda key: cells[col[key]].get_text(strip=True) if (key in col and col[key] < len(cells)) else ""
                code = cell("Subject Code")
                if not code:
                    continue
                tmarks = cell("Total Marks") or cell("Total") or cell("Marks")
                sub_rows.append({
                    "Subject Code": code,
                    "Subject Name": cell("Subject Name"),
                    "Internal Marks": cell("Internal Marks"),
                    "External Marks": cell("External Marks"),
                    "Total Marks": tmarks,
                    "Result": cell("Result"),
                    "Announced / Updated on": cell("Announced / Updated on"),
                })
                try:
                    total += int(tmarks)
                except:
                    pass
                max_total += 100

    pct = (total / max_total * 100) if max_total else 0

    return (usn, name), sub_rows, {
        "total_obtained": total,
        "total_max": max_total,
        "percentage": pct
    }


# -----------------------------------------
# Safe utilities for alert handling & refresh
# -----------------------------------------
def accept_alert_if_present(driver, timeout=0.5):
    """Attempt to accept an alert if present. Returns True if an alert was accepted."""
    try:
        alert = WebDriverWait(driver, timeout).until(EC.alert_is_present())
        try:
            print("[Info] Alert present — text:", alert.text)
        except Exception:
            pass
        alert.accept()
        return True
    except Exception:
        return False


def safe_refresh(driver):
    """Refresh the page while ensuring any modal alert is closed first."""
    # Try accepting any alert, then refresh
    try:
        try:
            # If an alert is present, accept it
            if accept_alert_if_present(driver, timeout=0.5):
                time.sleep(0.2)
        except Exception:
            pass
        driver.refresh()
    except UnexpectedAlertPresentException:
        # If an alert pops during refresh, accept and try again
        try:
            accept_alert_if_present(driver, timeout=0.5)
        except Exception:
            pass
        try:
            driver.refresh()
        except Exception:
            pass
    except WebDriverException as e:
        # Some WebDriver errors can be transient; print and continue
        print("[Warn] WebDriverException during refresh:", e)


# -----------------------------------------
# Data persistence: CSVs + live Excel rebuild
# -----------------------------------------
def save_interim(all_subs, all_summ):
    """Persist current progress to CSVs and rebuild the Excel in real time."""
    try:
        if all_subs or all_summ:
            pd.DataFrame(all_subs).to_csv(RAW_DATA, index=False)
            pd.DataFrame(all_summ).to_csv(RAW_SUMMARY, index=False)
            print(f"[Info] Saved progress: {len(all_subs)} subject rows, {len(all_summ)} student summaries.")
            generate_excel()
    except Exception as e:
        print("[Warn] Could not save interim CSVs:", e)


def _grade(marks):
    """Map total marks to a VTU grade (per the standard grade table)."""
    try:
        m = float(marks)
    except (TypeError, ValueError):
        return ""
    if m >= 90:
        return "O"
    if m >= 80:
        return "A+"
    if m >= 70:
        return "A"
    if m >= 60:
        return "B+"
    if m >= 55:
        return "B"
    if m >= 50:
        return "C"
    if m >= 40:
        return "P"
    return "F"


def generate_excel():
    """Rebuild vtu_results.xlsx from raw_results.csv + raw_summary.csv.
    Layout: one row per student; columns grouped per subject
    (Subject Name | Internal | External | Total | Grade | Result), then
    Percentage / Overall Total / Overall Max Marks. Header styled,
    panes frozen, sensible column widths.
    """
    try:
        if not os.path.exists(RAW_DATA) or not os.path.exists(RAW_SUMMARY):
            print("[Warn] Raw CSVs not found — Excel not generated yet.")
            return

        subs_df = pd.read_csv(RAW_DATA)
        summ_df = pd.read_csv(RAW_SUMMARY)

        if subs_df.empty or summ_df.empty:
            print("[Warn] No scraped data available — Excel not generated yet.")
            return

        subs_df.drop_duplicates(inplace=True)
        summ_df.drop_duplicates(subset=["USN"], keep="last", inplace=True)

        for df in (subs_df, summ_df):
            for col in ("USN", "Name"):
                if col in df.columns:
                    df[col] = df[col].astype(str).str.strip()

        codes = sorted(subs_df["Subject Code"].dropna().unique().tolist())
        sub_info = {}
        for _, r in subs_df.iterrows():
            key = (str(r["USN"]).strip(), r["Subject Code"])
            sub_info.setdefault(key, {})["name"] = str(r["Subject Name"])
            sub_info[key]["grade"] = _grade(r["Total Marks"])

        pivot = pd.pivot_table(
            subs_df,
            index=["USN", "Name"],
            columns="Subject Code",
            values=["Internal Marks", "External Marks", "Total Marks", "Result"],
            aggfunc="first"
        )
        pivot.columns = [f"{c2} - {c1}" for c1, c2 in pivot.columns]
        pivot.reset_index(inplace=True)

        # Reorder columns so every subject's block sits together:
        # USN | Name | [Subject Name | Internal | External | Total | Grade | Result] per subject | Percentage | Overall Total | Overall Max Marks
        order = ["USN", "Name"]
        for code in codes:
            order.append(f"{code} - Subject Name")
            for suf in ("Internal Marks", "External Marks", "Total Marks", "Grade", "Result"):
                order.append(f"{code} - {suf}")
        for col in order:
            if col not in pivot.columns:
                pivot[col] = ""
        pivot = pivot[order].copy()

        pivot["USN"] = pivot["USN"].astype(str)
        pivot["Name"] = pivot["Name"].astype(str)
        for idx in pivot.index:
            subkey = str(pivot.at[idx, "USN"])
            for code in codes:
                info = sub_info.get((subkey, code))
                if info:
                    pivot.at[idx, f"{code} - Subject Name"] = info.get("name", "")
                    pivot.at[idx, f"{code} - Grade"] = info.get("grade", "")
        pivot = pivot.replace({pd.NA: ""}).fillna("")

        summ_df["Percentage"] = summ_df["percentage"].apply(lambda x: f"{x:.2f}%" if isinstance(x, (int, float)) else "")
        summ_df = summ_df.rename(columns={
            "total_obtained": "Overall Total",
            "total_max": "Overall Max Marks"
        })

        final = pd.merge(pivot, summ_df, on=["USN", "Name"], how="outer")
        final = final.replace({pd.NA: ""}).fillna("")
        final = final.drop_duplicates(subset=["USN", "Name"], keep="last")
        final.to_excel(OUTPUT_EXCEL, index=False, startrow=HEADER_ROWS)

        _style_excel(OUTPUT_EXCEL, final.columns.tolist())
        print("[Success] Excel updated:", OUTPUT_EXCEL)

    except pd.errors.EmptyDataError:
        print("[Warn] Raw CSVs are empty — Excel not generated yet.")
    except PermissionError:
        print("[Error] Cannot write vtu_results.xlsx — it is open in Excel. Close it and re-run the fetch.")
    except Exception as e:
        print("[Error] Excel generation failed:", e)
        traceback.print_exc()


def _style_excel(path, columns):
    """Apply a clean look: college header block, bold header on a green fill,
    frozen panes, per-column widths and an autofilter."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(path)
        ws = wb.active

        head_fill = PatternFill("solid", fgColor="1C7A5E")
        head_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="C8D6CE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[DATA_HEADER_ROW]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        widths = {"USN": 14, "Name": 28}
        for i, col in enumerate(columns, start=1):
            base = "USN" if col == "USN" else "Name" if col == "Name" else "subj" if "Subject Name" in col else "mark" if col in ("Percentage", "Overall Total", "Overall Max Marks") else "num"
            ws.column_dimensions[ws.cell(row=DATA_HEADER_ROW, column=i).column_letter].width = widths.get(base, 12 if base == "num" else 34 if base == "subj" else 14)
        for row in ws.iter_rows(min_row=DATA_HEADER_ROW + 1):
            for cell in row:
                cell.border = border

        apply_header(ws, len(columns))
        ws.freeze_panes = f"C{DATA_HEADER_ROW + 1}"
        ws.auto_filter.ref = f"A{DATA_HEADER_ROW}:{get_column_letter(len(columns))}{ws.max_row}"
        wb.save(path)
    except Exception as e:
        print("[Warn] Excel styling skipped:", e)


# -----------------------------------------
# Browser management (with auto-restart on crash)
# -----------------------------------------
def start_driver():
    """Create a fresh Chrome WebDriver."""
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--log-level=3")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def browser_alive(driver):
    """Check whether the browser session is still usable."""
    try:
        driver.current_url
        return True
    except Exception:
        return False


def ensure_browser(driver):
    """Restart the browser if it crashed (window closed, session dead, etc.)."""
    if browser_alive(driver):
        return driver
    print("[Warn] Browser session is dead — restarting Chrome...")
    try:
        driver.quit()
    except Exception:
        pass
    return start_driver()


# -----------------------------------------
# MAIN
# -----------------------------------------
def main():
    try:
        df = pd.read_csv(INPUT_CSV)
        usns = df["USN"].astype(str).str.strip().tolist()
    except Exception as e:
        print("Error reading students.csv:", e)
        return

    # Remove stale outputs from previous runs so old data never leaks into the new fetch
    for f in (RAW_DATA, RAW_SUMMARY, OUTPUT_EXCEL):
        try:
            os.remove(f)
            print(f"[Info] Cleared stale file: {f}")
        except FileNotFoundError:
            pass
        except PermissionError:
            print(f"[Warn] Could not clear {f} — is it open in Excel? Close it and re-run the fetch.")

    try:
        model = tf.keras.models.load_model(MODEL_FILE)
    except Exception as e:
        print("Error loading CAPTCHA model:", e)
        return

    # Start Chrome WebDriver with automatic driver management
    try:
        driver = start_driver()
    except Exception as e:
        print("Error starting Chrome WebDriver:", e)
        return

    all_subs = []
    all_summ = []
    unique = set()

    for usn in usns:
        print(f"\n--- Processing USN: {usn} ---")

        driver = ensure_browser(driver)

        try:
            driver.get(VTU_RESULTS_URL)
        except Exception as e:
            print("[Error] driver.get failed:", e)
            safe_refresh(driver)

        wait = WebDriverWait(driver, 10)
        success = False

        for attempt in range(1, MAX_ATTEMPTS + 1):
            print(f"Attempt {attempt}/{MAX_ATTEMPTS} for {usn}")
            try:
                # Wait for the form fields
                box = wait.until(EC.presence_of_element_located((By.NAME, "lns")))
                cbox = driver.find_element(By.NAME, "captchacode")
                img = driver.find_element(By.XPATH, "//img[contains(@src,'vtu_captcha.php')]")
                btn = driver.find_element(By.ID, "submit")

                # --- Step 1: solve the CAPTCHA first (before filling anything) ---
                img.screenshot("cap.png")

                # Preprocess & predict
                try:
                    prep = preprocess("cap.png")
                    pred = decode(predict_captcha(model, prep))
                except Exception as e:
                    print("[Warn] CAPTCHA preprocess/predict failed:", e)
                    pred = ""

                print("CAPTCHA predicted:", pred)

                # If prediction length not 6 -> safe refresh and retry
                if not isinstance(pred, str) or len(pred) != 6:
                    print("[Info] Pred length invalid, refreshing and retrying.")
                    safe_refresh(driver)
                    time.sleep(0.2)
                    continue

                # --- Step 2: now fill the USN and the solved CAPTCHA ---
                try:
                    cbox.clear()
                except Exception:
                    pass
                try:
                    box.clear()
                except Exception:
                    pass
                box.send_keys(usn)

                # Enter captcha and submit
                try:
                    cbox.send_keys(pred)
                except Exception:
                    try:
                        cbox.clear()
                        cbox.send_keys(pred)
                    except Exception:
                        pass

                # Store current url to detect change
                try:
                    old_url = driver.current_url
                except UnexpectedAlertPresentException:
                    # If an alert is present while getting URL, accept it and refresh
                    accept_alert_if_present(driver, timeout=0.5)
                    safe_refresh(driver)
                    continue

                # Click submit safely
                try:
                    btn.click()
                except UnexpectedAlertPresentException:
                    # Alert popped during click; accept and retry refresh
                    accept_alert_if_present(driver, timeout=0.5)
                    safe_refresh(driver)
                    time.sleep(0.5)
                    continue
                except Exception as e:
                    print("[Warn] Click failed:", e)
                    # try JavaScript click as fallback
                    try:
                        driver.execute_script("arguments[0].click();", btn)
                    except Exception:
                        safe_refresh(driver)
                        continue

                # Immediately check for alert (invalid captcha)
                if accept_alert_if_present(driver, timeout=1):
                    print("[Info] Detected alert after submit (likely invalid captcha). Retrying.")
                    safe_refresh(driver)
                    time.sleep(0.5)
                    continue

                # Wait for page navigation (url change) — if alert occurs here, handle it
                try:
                    WebDriverWait(driver, 4).until(EC.url_changes(old_url))
                except UnexpectedAlertPresentException:
                    # If an unexpected alert interrupts, accept and retry
                    accept_alert_if_present(driver, timeout=0.5)
                    safe_refresh(driver)
                    time.sleep(0.5)
                    continue
                except TimeoutException:
                    # URL did not change — might be invalid captcha or same page loaded
                    # Check for alert one more time
                    if accept_alert_if_present(driver, timeout=1):
                        safe_refresh(driver)
                        time.sleep(0.5)
                        continue
                    # No alert and no url change — retry
                    safe_refresh(driver)
                    time.sleep(0.5)
                    continue

                # If we reached here, page changed successfully — wait until the
                # result content is actually present in the DOM, then scrape it
                try:
                    WebDriverWait(driver, 10).until(
                        lambda d: ("divTableBody" in d.page_source) or ("table-condensed" in d.page_source)
                    )
                except TimeoutException:
                    pass  # handled below by the 0-subjects retry

                try:
                    page_html = driver.page_source
                    (u, name), subs, summ = scrape(page_html)

                    # Page loaded but no result content found — could be a slow
                    # render, changed layout, or "result not found" page.
                    if len(subs) == 0:
                        try:
                            with open("debug_page.html", "w", encoding="utf-8") as df:
                                df.write(page_html)
                            print("[Debug] 0 subjects scraped — page saved to debug_page.html (URL: %s)" % driver.current_url)
                        except Exception:
                            pass
                        if "not found" in page_html.lower() or "no result" in page_html.lower() or "not declared" in page_html.lower():
                            print(f"[Info] Result not found for {usn} on VTU site.")
                            success = False
                            break
                        print("[Warn] Scrape returned 0 subjects — refreshing and retrying.")
                        safe_refresh(driver)
                        time.sleep(0.5)
                        continue

                    for s in subs:
                        s["USN"] = u
                        s["Name"] = name
                        all_subs.append(s)
                        unique.add(s["Subject Code"])

                    summ["USN"] = u
                    summ["Name"] = name
                    all_summ.append(summ)

                    success = True
                    print(f"[Success] Scraped {u} - {name} with {len(subs)} subjects.")
                    break

                except Exception as e:
                    print("[Error] Scrape failed:", e)
                    traceback.print_exc()
                    safe_refresh(driver)
                    time.sleep(0.5)
                    continue

            except UnexpectedAlertPresentException:
                # Always accept unexpected alerts and retry
                try:
                    accept_alert_if_present(driver, timeout=0.5)
                except Exception:
                    pass
                safe_refresh(driver)
                time.sleep(0.5)
                continue

            except TimeoutException:
                print("[Warn] Timeout waiting for page elements. Refreshing and retrying.")
                driver = ensure_browser(driver)
                safe_refresh(driver)
                time.sleep(0.5)
                continue

            except Exception as e:
                print("[Error] Unexpected exception in attempt loop:", e)
                traceback.print_exc()
                driver = ensure_browser(driver)
                safe_refresh(driver)
                time.sleep(0.5)
                continue

        if not success:
            print(f"FAILED to fetch results for USN: {usn}")
            # Add a summary row marking the failure if desired
            all_summ.append({'USN': usn, 'Name': 'FETCH FAILED', 'percentage': 0, 'total_obtained': 0, 'total_max': 0})

        # Save intermediate CSVs so partial progress is kept (Excel rebuilt live)
        save_interim(all_subs, all_summ)

    # End for all USNs
    try:
        driver.quit()
    except Exception:
        pass

    print("\n--- SCRAPING COMPLETE ---")
    print("[SUBJECTS]:" + ",".join(sorted(unique)) + f" (run {RUN_ID})")

    # Generate/refresh Excel without SGPA (SGPA added by app.py)
    generate_excel()


if __name__ == "__main__":
    main()
