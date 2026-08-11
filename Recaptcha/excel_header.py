"""Header block for the result Excel files (shared by the live fetch and the
batch-export paths).

Layout: row 1 = the header_banner.jpeg image (centered), row 2 = spacer,
data starts at row 3. Nothing else.
"""

import os

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BANNER_IMAGE = os.path.join(BASE_DIR, "header_banner.jpeg")

HEADER_ROWS = 2
DATA_HEADER_ROW = HEADER_ROWS + 1  # row 3

BANNER_HEIGHT_PX = 78
BANNER_Y_OFFSET_PX = 6
ROW_HEIGHTS = [70, 8]  # rows 1-2


def apply_header(ws, last_col):
    """Write the banner image into row 1, centered across the table width.
    last_col = number of data columns (int)."""
    try:
        ws.row_dimensions[1].height = ROW_HEIGHTS[0]
        ws.row_dimensions[2].height = ROW_HEIGHTS[1]

        _blank_header_cells(ws, last_col)
        _anchor_banner(ws, last_col)
    except Exception as e:
        print("[Warn] Header skipped:", e)


def _blank_header_cells(ws, last_col):
    """Solid white fill over the header rows (1..HEADER_ROWS). Gridlines stay
    ON for the rest of the sheet, but Excel hides them behind a filled cell,
    so the header area renders as a clean white band while the result table
    below keeps its normal gridlines and borders."""
    try:
        white = PatternFill("solid", fgColor="FFFFFF")
        for r in range(1, HEADER_ROWS + 1):
            for c in range(1, max(last_col, 2) + 1):
                ws.cell(row=r, column=c).fill = white
    except Exception as e:
        print("[Warn] Header cell blanking skipped:", e)


def _anchor_banner(ws, last_col):
    """Place the wide banner image in row 1, horizontally centered across the
    full table width, using an explicit one-cell anchor with pixel offset."""
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D

        if not os.path.exists(BANNER_IMAGE):
            return

        img = XLImage(BANNER_IMAGE)
        img.height = BANNER_HEIGHT_PX
        img.width = int(BANNER_HEIGHT_PX * _aspect(BANNER_IMAGE))

        total_px = 0
        for c in range(1, max(last_col, 2) + 1):
            w = ws.column_dimensions[get_column_letter(c)].width
            total_px += int((w if w else 8.43) * 7 + 5)

        off_px = max(0, int((total_px - img.width) / 2))
        marker = AnchorMarker(
            col=0, colOff=pixels_to_EMU(off_px), row=0, rowOff=pixels_to_EMU(BANNER_Y_OFFSET_PX)
        )
        img.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(
                cx=pixels_to_EMU(img.width),
                cy=pixels_to_EMU(img.height),
            ),
        )
        ws.add_image(img)
    except Exception as e:
        print("[Warn] Banner embedding skipped:", e)


def _aspect(path):
    try:
        from PIL import Image as PILImage

        with PILImage.open(path) as im:
            w, h = im.size
            return (w / h) if h else 1.0
    except Exception:
        return 1.0