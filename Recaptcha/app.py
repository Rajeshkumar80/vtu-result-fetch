import io
import os
import re
import sys
import time
import subprocess
import pandas as pd
from datetime import datetime
from flask import Flask, render_template, send_from_directory, send_file, make_response, jsonify
from flask_socketio import SocketIO, emit
import db
from excel_header import HEADER_ROWS, DATA_HEADER_ROW, apply_header

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BACKEND_SCRIPT = os.path.join(BASE_DIR, "bulk_fetcher_6.py")
DEFAULT_CSV = os.path.join(BASE_DIR, "students.csv")
RAW_DATA = os.path.join(BASE_DIR, "raw_results.csv")
RAW_SUMMARY = os.path.join(BASE_DIR, "raw_summary.csv")
OUTPUT_EXCEL = os.path.join(BASE_DIR, "vtu_results.xlsx")
PY = sys.executable
USN_PATTERN = re.compile(r"^[A-Z0-9]{10}$")

current_process = None
fetch_stats = {"total": 0, "success": 0, "fail": 0}
current_run_id = None

app = Flask(__name__)
socketio = SocketIO(app, async_mode="threading", cors_allowed_origins="*")


# -----------------------------------------
# FRONTEND ROUTES
# -----------------------------------------
@app.route("/")
def index():
    resp = make_response(render_template("index.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route("/download")
def download():
    """Serve the latest vtu_results.xlsx with caching fully disabled so a
    browser can never return a stale copy of a previous run's file."""
    try:
        resp = send_from_directory(
            BASE_DIR, os.path.basename(OUTPUT_EXCEL),
            as_attachment=True, conditional=False
        )
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp
    except FileNotFoundError:
        return "Excel not generated yet", 404


@app.route("/export-batch/<batch_id>")
def export_batch(batch_id):
    """Regenerate an .xlsx from MongoDB data for a saved batch (independent of
    the live vtu_results.xlsx file)."""
    batch, students = db.fetch_batch_with_students(batch_id)
    if batch is None or students is None:
        return "Batch not found", 404

    df = students_to_pivot_df(students)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=HEADER_ROWS)
    _style_excel_buffer(buf, df.columns.tolist())
    buf.seek(0)

    fname = f"vtu_results_{batch.get('department','')}_{batch.get('semester','')}_{batch.get('scheme','')}_{batch.get('year','')}.xlsx"
    resp = send_file(
        buf, as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        conditional=False
    )
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp


def _style_excel_buffer(buf, columns):
    """Apply a clean look to an xlsx already written into buf: college header
    block, bold header on a green fill, frozen panes, per-column widths and
    an autofilter."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active

        head_fill = PatternFill("solid", fgColor="1C7A5E")
        head_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="C8D6CE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[DATA_HEADER_ROW]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for i, col in enumerate(columns, start=1):
            letter = ws.cell(row=DATA_HEADER_ROW, column=i).column_letter
            if col == "USN":
                ws.column_dimensions[letter].width = 14
            elif col == "Name":
                ws.column_dimensions[letter].width = 28
            elif "Subject Name" in col:
                ws.column_dimensions[letter].width = 34
            elif "Result Status" in col or "Percentage" in col:
                ws.column_dimensions[letter].width = 15
            else:
                ws.column_dimensions[letter].width = 12

        for row in ws.iter_rows(min_row=DATA_HEADER_ROW + 1):
            for cell in row:
                cell.border = border

        apply_header(ws, len(columns))
        ws.freeze_panes = f"C{DATA_HEADER_ROW + 1}"
        ws.auto_filter.ref = f"A{DATA_HEADER_ROW}:{get_column_letter(len(columns))}{ws.max_row}"
        buf.seek(0)
        buf.truncate()
        wb.save(buf)
    except Exception as e:
        print("[Warn] Excel styling skipped:", e)


def _grade(marks):
    """Map total marks to a VTU grade (per the standard grade table)."""
    try:
        m = float(marks)
    except (TypeError, ValueError):
        return ""
    if m >= 90:
        return "O"
    if m >= 80:
        return "A+"
    if m >= 70:
        return "A"
    if m >= 60:
        return "B+"
    if m >= 55:
        return "B"
    if m >= 50:
        return "C"
    if m >= 40:
        return "P"
    return "F"


def students_to_pivot_df(students):
    """Turn DB student docs into the same pivot shape as the live Excel."""
    rows = []
    for s in students:
        row = {
            "USN": s.get("usn", ""),
            "Name": s.get("name", ""),
            "Percentage": s.get("percentage"),
            "Result Status": s.get("result_status", ""),
        }
        for sub in s.get("subjects", []):
            code = sub.get("code", "")
            row[f"{code} - Internal Marks"] = sub.get("internal")
            row[f"{code} - External Marks"] = sub.get("external")
            row[f"{code} - Total Marks"] = sub.get("total")
            row[f"{code} - Grade"] = _grade(sub.get("total"))
            row[f"{code} - Result"] = sub.get("result")
        rows.append(row)
    return pd.DataFrame(rows)


# -----------------------------------------
# FILTER OPTIONS API (populates Browse dropdowns from the database)
# -----------------------------------------
SEMESTER_PRESETS = [str(s) for s in range(1, 9)]
SCHEME_PRESETS = [str(y) for y in range(2018, 2027)]
YEAR_PRESETS = [str(y) for y in range(2019, 2027)]


def _num_key(v):
    return int(v) if v.isdigit() else float("inf")


def _filter_values(field, presets, reverse=False):
    """Merge static presets with distinct DB values; dedupe; numeric sort."""
    try:
        merged = set(presets) | set(db.distinct_batch_values(field))
    except Exception:
        merged = set(presets)
    return sorted(merged, key=_num_key, reverse=reverse)


@app.route("/api/filters/semesters")
def filters_semesters():
    return jsonify(_filter_values("semester", SEMESTER_PRESETS))


@app.route("/api/student-record/<usn>")
def student_record(usn):
    records, err = db.find_student_records(usn)
    if err:
        return jsonify({"error": err}), 404
    name = ""
    for r in records:
        if r.get("name"):
            name = r["name"]
            break
    return jsonify({"name": name, "records": records})


@app.route("/api/filters/schemes")
def filters_schemes():
    return jsonify(_filter_values("scheme", SCHEME_PRESETS, reverse=True))


@app.route("/api/filters/years")
def filters_years():
    return jsonify(_filter_values("year", YEAR_PRESETS))


# -----------------------------------------
# IMPORT CSV
# -----------------------------------------
@socketio.on("import-csv")
def import_csv():
    if not os.path.exists(DEFAULT_CSV):
        emit("log-message", {"data": "[Error] students.csv missing\n"})
        return

    with open(DEFAULT_CSV) as f:
        lines = f.readlines()[1:]

    emit("csv-data", {"usns": "".join(lines)})
    emit("log-message", {"data": f"Imported {len(lines)} USNs\n"})


# -----------------------------------------
# START FETCHING
# -----------------------------------------
@socketio.on("start-fetch")
def start_fetch(msg):
    global fetch_stats, current_process, current_run_id

    usns = msg.get("usns", [])
    vtu_url = msg.get("url", "")

    if current_process and current_process.poll() is None:
        emit("log-message", {"data": "[Error] A fetch is already running. Stop it first.\n"})
        return

    if not usns:
        emit("log-message", {"data": "[Error] No USNs provided.\n"})
        return

    if not vtu_url:
        emit("log-message", {"data": "[Error] No VTU URL provided.\n"})
        return

    # Validate & clean USNs server-side
    clean = []
    skipped = []
    for u in usns:
        u = str(u).strip().upper()
        if USN_PATTERN.match(u):
            clean.append(u)
        else:
            skipped.append(u)

    if skipped:
        emit("log-message", {"data": f"[Warn] Skipped invalid USN entries: {', '.join(skipped)}\n"})
    if not clean:
        emit("log-message", {"data": "[Error] No valid USNs. Format e.g. 1GD23CS001\n"})
        return

    with open(DEFAULT_CSV, "w") as f:
        f.write("USN\n")
        for u in clean:
            f.write(u + "\n")

    fetch_stats = {"total": len(clean), "success": 0, "fail": 0}
    current_run_id = time.strftime("%Y%m%d-%H%M%S")

    emit("log-message", {"data": f"[Run {current_run_id}] Fetch started — {len(clean)} USN(s).\n"})

    # Clear stale outputs from previous runs so old data never leaks into the new fetch
    for f in (RAW_DATA, RAW_SUMMARY, OUTPUT_EXCEL):
        try:
            os.remove(f)
        except FileNotFoundError:
            pass
        except PermissionError:
            emit("log-message", {"data": f"[Warn] Could not clear {os.path.basename(f)} — close it if it is open in Excel.\n"})

    socketio.start_background_task(target=run_scraper, vtu_url=vtu_url, total_usns=len(clean), run_id=current_run_id)
    emit("fetch-started", {"total": len(clean), "run_id": current_run_id})


# -----------------------------------------
# STOP FETCHING
# -----------------------------------------
@socketio.on("stop-fetch")
def stop_fetch():
    global current_process
    if current_process and current_process.poll() is None:
        current_process.terminate()
        emit("log-message", {"data": "\n[Stopped] Fetch terminated by user.\n"})
    else:
        emit("log-message", {"data": "\n[Info] No running process to stop.\n"})


# -----------------------------------------
# RUN SCRAPER WITH URL ARG
# -----------------------------------------
def run_scraper(vtu_url, total_usns, run_id):
    global current_process, fetch_stats
    cmd = [PY, BACKEND_SCRIPT, vtu_url, run_id]
    process = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        cwd=BASE_DIR
    )
    current_process = process
    started = time.time()
    failed_usns = []

    for line in iter(process.stdout.readline, ""):
        socketio.emit("log-message", {"data": line})
        if "[Success] Scraped" in line:
            fetch_stats["success"] += 1
            socketio.emit("fetch-progress", dict(fetch_stats))
        elif "FAILED to fetch" in line:
            fetch_stats["fail"] += 1
            m = re.search(r"FAILED to fetch results for USN: (\S+)", line)
            if m:
                failed_usns.append(m.group(1))
            socketio.emit("fetch-progress", dict(fetch_stats))

    process.wait()
    current_process = None

    complete_data = dict(fetch_stats)
    complete_data["failed"] = failed_usns
    socketio.emit("fetch-complete", complete_data)

    # Explicit write confirmation: only advertise the download if the Excel was
    # actually rewritten during this run (mtime newer than when we started).
    fresh = False
    if os.path.exists(OUTPUT_EXCEL):
        try:
            fresh = os.path.getmtime(OUTPUT_EXCEL) >= started - 1
        except OSError:
            fresh = False

    if fresh:
        mtime = os.path.getmtime(OUTPUT_EXCEL)
        socketio.emit("log-message", {"data": f"[Run {run_id}] Excel verified fresh (mtime {time.strftime('%H:%M:%S', time.localtime(mtime))}).\n"})
        socketio.emit("download-ready")
    else:
        socketio.emit("log-message", {"data": f"\n[Run {run_id}] [Error] No fresh results were generated — check the logs above (vtu_results.xlsx may be open in Excel, or all USNs failed).\n"})


# -----------------------------------------
# MONGODB EVENTS
# -----------------------------------------
@socketio.on("get-db-status")
def get_db_status():
    emit("db-status", {"connected": db.is_connected(), "message": db.status_msg})


def _to_num(v):
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return None


def build_db_payload(year, scheme, semester, department):
    """Read the just-fetched raw CSVs and build (batch_doc, student_docs).
    Returns (None, error_message) on any failure."""
    if not os.path.exists(RAW_DATA) or not os.path.exists(RAW_SUMMARY):
        return None, "No fetched data found. Run a fetch first."

    try:
        subs = pd.read_csv(RAW_DATA)
        summ = pd.read_csv(RAW_SUMMARY)
    except Exception as e:
        return None, f"Could not read raw CSVs: {e}"

    if subs.empty or summ.empty:
        return None, "Raw CSVs are empty — nothing to save."

    subs["USN"] = subs["USN"].astype(str).str.strip().str.upper()
    summ["USN"] = summ["USN"].astype(str).str.strip().str.upper()
    summ_map = {r["USN"]: r for _, r in summ.iterrows()}

    subject_codes = sorted(subs["Subject Code"].dropna().unique().tolist())
    usns = sorted(subs["USN"].dropna().unique().tolist())
    prefix = re.sub(r"\d+$", "", usns[0]) if usns else ""

    student_docs = []
    for usn in usns:
        group = subs[subs["USN"] == usn]
        name = ""
        if not group.empty:
            name = str(group.iloc[0].get("Name", "") or "").strip()
        srow = summ_map.get(usn)
        percentage = None
        if srow is not None and pd.notna(srow.get("percentage")):
            percentage = round(float(srow["percentage"]), 2)

        subjects = []
        has_fail = False
        for _, r in group.iterrows():
            result = str(r.get("Result", "") or "").strip()
            if result == "F":
                has_fail = True
            subjects.append({
                "code": str(r["Subject Code"]).strip(),
                "subject_name": str(r.get("Subject Name", "") or "").strip(),
                "internal": _to_num(r.get("Internal Marks")),
                "external": _to_num(r.get("External Marks")),
                "total": _to_num(r.get("Total Marks")),
                "result": result,
            })

        student_docs.append({
            "usn": usn,
            "name": name,
            "subjects": subjects,
            "percentage": percentage,
            "result_status": "FAIL" if has_fail else "PASS",
        })

    batch_doc = {
        "year": str(year).strip(),
        "scheme": str(scheme).strip(),
        "semester": str(semester).strip(),
        "department": str(department).strip(),
        "saved_at": datetime.utcnow(),
        "usn_prefix": prefix,
        "student_count": len(student_docs),
        "subjects": subject_codes,
        "run_id": current_run_id,
    }
    return batch_doc, student_docs


@socketio.on("save-to-db")
def save_to_db(data):
    year = str(data.get("year", "")).strip()
    scheme = str(data.get("scheme", "")).strip()
    semester = str(data.get("semester", "")).strip()
    department = str(data.get("department", "")).strip()
    target_id = str(data.get("target_batch_id", "") or "").strip()

    if not db.is_connected():
        emit("log-message", {"data": f"[DB ERROR] MongoDB not available: {db.status_msg}\n"})
        return

    if target_id:
        batch = db.fetch_batch(target_id)
        if batch is None:
            emit("log-message", {"data": f"[DB ERROR] Target batch {target_id} not found.\n"})
            return
        year = str(batch.get("year", "") or "").strip()
        scheme = str(batch.get("scheme", "") or "").strip()
        semester = str(batch.get("semester", "") or "").strip()
        department = str(batch.get("department", "") or "").strip()

    if not (year and scheme and semester and department):
        emit("log-message", {"data": "[DB ERROR] year, scheme, semester and department are all required.\n"})
        return

    batch_doc, student_docs = build_db_payload(year, scheme, semester, department)
    if batch_doc is None:
        emit("log-message", {"data": f"[DB ERROR] {student_docs}\n"})
        return

    if target_id:
        batch_id, added, merged, total = db.merge_into_batch(target_id, batch_doc, student_docs)
    else:
        batch_id, added, merged, total = db.save_batch(batch_doc, student_docs)

    if batch_id is None:
        emit("log-message", {"data": f"[DB ERROR] Save failed: {added}\n"})
        return

    if merged or target_id:
        emit("log-message", {
            "data": f"[DB] Merged into batch {batch_id} — {added} new student(s), total {total}.\n"
        })
    else:
        emit("log-message", {
            "data": f"[DB] Saved new batch {batch_id} — {total} students, {len(batch_doc['subjects'])} subjects.\n"
        })
    emit("save-to-db-complete", {
        "batch_id": str(batch_id),
        "student_count": total,
        "added_count": added,
        "merged": merged,
    })


@socketio.on("get-batches")
def get_batches():
    if not db.is_connected():
        emit("batches", {"batches": [], "error": db.status_msg})
        return
    batches = db.fetch_batches()
    emit("batches", {"batches": batches, "error": None})


@socketio.on("get-batch-results")
def get_batch_results(data):
    batch_id = str(data.get("batch_id", ""))
    if not db.is_connected():
        emit("batch-results", {"batch": None, "students": [], "error": db.status_msg})
        return
    batch, students = db.fetch_batch_with_students(batch_id)
    emit("batch-results", {"batch": batch, "students": students, "error": None})


# -----------------------------------------
if __name__ == "__main__":
    db.init_db()
    print(f"[MongoDB] status={db.status} — {db.status_msg}")
    socketio.run(app, host="127.0.0.1", port=5000, allow_unsafe_werkzeug=True, use_reloader=False)
